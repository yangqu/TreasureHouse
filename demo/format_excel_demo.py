# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 10:51:06 2020

@author: Focusmedia
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Color, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Alignment, Border, Side
import numpy as np
from openpyxl.formatting.rule import DataBar, Rule, FormulaRule, FormatObject
from openpyxl.utils import get_column_letter, column_index_from_string
import os

'''
inputdir=input('需要处理的文件路径:')
inputname=input('需要处理的文件名:')
outputdir=input('需要输出的路径名:')
'''

inputdir = sys.argv[1]
inputname = sys.argv[2]
date_file = sys.argv[3]
#inputdir = r"D:/data/excel/"
#inputname = "sale.txt"
#date_file = "20200717"
date_file_output = inputname.split('.')[0]


def monitor_result_sheet10(inputname, inputdir, outputdir):

    # 确定结果sheet名
    result_sheetname = '项目推送率' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    sheet10_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == '全媒体') & (impt_xlsx_resource.套装 == '不分套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['项目总数', '成功项目总数', '项目推送成功率'], aggfunc=np.sum)

    sheet10_dataframe = sheet10_dataframe.sort_values('项目总数', ascending=False)

    # 重置列的顺序
    order = ['项目总数', '成功项目总数', '项目推送成功率']
    sheet10_dataframe = sheet10_dataframe[order]

    ##返回行数
    row_1 = sheet10_dataframe.shape[0]
    ##返回列数
    column_1 = sheet10_dataframe.shape[1] - 1
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet10_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=2)
    writer.save()
    writer.close()

    df = sheet10_dataframe.reset_index('城市')
    df1 = df.loc[(df['城市'] != '全国') & (df['项目推送成功率'] > 0.5)]
    df2 = df.loc[(df['城市'] != '全国') & (df['项目推送成功率'] > 0.7)]
    # print(df)
    # print("%.f%%" % (df.iat[0, 3] * 100))
    # print(df1)
    # print(df2)
    # print(len(df1))
    # print(len(df2))

    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 9
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    # 获取长度时为了让中文识别为两个字符，需要控制encode为u8
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 14
    sheet.column_dimensions['D'].width = 14

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column == 4:
                    cell.number_format = number_format
    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = '项目推送成功率'
    sheet['A1'].font = fonta1

    sheet['A1'].alignment = Alignment(horizontal='centerContinuous', wrapText=True)
    sheet.merge_cells('A1:D1')
    sheet['A2'] = '''项目推送情况说明-''' + date_file + '''：
    1、本周项目推送成功率为''' + str("%.f%%" % (df.iat[0, 3] * 100)) + '''
    2、本周全国''' + str(len(df1)) + '''个城市的项目推送成功率超过50%，''' + str(len(df2)) + '''个城市超过70%。'''
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    # 当通过指定单元格写入带有换行符的文字时，为了让文件打开时显示的就是已换行的，需要对Aligmentj中的wrapText属性设置为True

    sheet.merge_cells('A2:D2')
    sheet['A2'].alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 75
    sheet.row_dimensions[1].height = 21

    # 设置条件格式
    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[0.7])
    green_fill = PatternFill(bgColor='C0F0D8')
    green_text = Font(color='006000')
    dxf_between = DifferentialStyle(font=green_text, fill=green_fill)
    rule2 = Rule(type='cellIs', operator='between', dxf=dxf_between, formula=[0.5, 0.7])
    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 1, maxrow), rule1)
    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 1, maxrow), rule2)

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet9(inputname, inputdir, outputdir):
    # 获取文件名

    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = 'LCD项目推送率' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    sheet10_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'lcd') & (impt_xlsx_resource.套装 == '不分套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['项目总数', '成功项目总数', '项目推送成功率'], aggfunc=np.sum)

    sheet10_dataframe = sheet10_dataframe.sort_values('项目总数', ascending=False)

    # 重置列的顺序
    order = ['项目总数', '成功项目总数', '项目推送成功率']
    sheet10_dataframe = sheet10_dataframe[order]

    ##返回行数
    row_1 = sheet10_dataframe.shape[0]
    ##返回列数
    column_1 = sheet10_dataframe.shape[1] - 1
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet10_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=2)
    writer.save()
    writer.close()
    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 8
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 14
    sheet.column_dimensions['D'].width = 14

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column == 4:
                    cell.number_format = number_format

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = 'LCD项目推送成功率'
    sheet['A1'].font = fonta1

    sheet['A1'].alignment = Alignment(horizontal='centerContinuous', wrapText=True)
    sheet.merge_cells('A1:D1')

    df = sheet10_dataframe
    df = df[df.index.values != '全国']
    city = df[df['项目推送成功率'] > 0.7]
    if len(city) > 4:
        city = '、'.join(city[0:4].index)
    else:
        city = '、'.join(city.index)

    sheet['A2'] = 'LCD项目推送情况说明-' + date_file + '：\n1、本周项目推送成功率为' + str(
        '%.1f' % (round(sheet['D4'].value * 100, 1))) + '%\n2、本周全国' + str(
        np.sum(df['项目推送成功率'] > 0.5)) + '个城市的项目推送成功率超过50%，' + city + '等' + str(np.sum(df['项目推送成功率'] > 0.7)) + '个城市超过70%.'
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    # 当通过指定单元格写入带有换行符的文字时，为了让文件打开时显示的就是已换行的，需要对Aligmentj中的wrapText属性设置为True

    sheet.merge_cells('A2:D2')
    sheet['A2'].alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 75
    sheet.row_dimensions[1].height = 21

    # 设置条件格式
    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[0.7])
    green_fill = PatternFill(bgColor='C0F0D8')
    green_text = Font(color='006000')
    dxf_between = DifferentialStyle(font=green_text, fill=green_fill)
    rule2 = Rule(type='cellIs', operator='between', dxf=dxf_between, formula=[0.5, 0.7])
    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 1, maxrow), rule1)
    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 1, maxrow), rule2)

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet8(inputname, inputdir, outputdir):
    # 获取文件名

    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = '智能屏项目推送率' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    sheet10_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'smart') & (impt_xlsx_resource.套装 == '不分套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['项目总数', '成功项目总数', '项目推送成功率'], aggfunc=np.sum)
    sheet10_dataframe = sheet10_dataframe.sort_values('项目总数', ascending=False)

    # 重置列的顺序
    order = ['项目总数', '成功项目总数', '项目推送成功率']
    sheet10_dataframe = sheet10_dataframe[order]

    ##返回行数
    row_1 = sheet10_dataframe.shape[0]
    ##返回列数
    column_1 = sheet10_dataframe.shape[1] - 1
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet10_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=2)
    writer.save()
    writer.close()
    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 7
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 14
    sheet.column_dimensions['D'].width = 14

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column == 4:
                    cell.number_format = number_format
    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = '项目推送成功率'
    sheet['A1'].font = fonta1

    sheet['A1'].alignment = Alignment(horizontal='centerContinuous', wrapText=True)
    sheet.merge_cells('A1:D1')

    df = sheet10_dataframe
    df = df[df.index.values != '全国']

    sheet['A2'] = '智能屏项目推送情况说明-' + date_file + '：\n1、本周项目推送成功率为' + str(
        '%.1f' % (round(sheet['D4'].value * 100, 1))) + '%\n2、本周全国' + str(
        np.sum(df['项目推送成功率'] > 0.5)) + '个城市的项目推送成功率超过50%，' + str(np.sum(df['项目推送成功率'] > 0.7)) + '个城市超过70%。'
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    # 当通过指定单元格写入带有换行符的文字时，为了让文件打开时显示的就是已换行的，需要对Aligmentj中的wrapText属性设置为True

    sheet.merge_cells('A2:D2')
    sheet['A2'].alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 75
    sheet.row_dimensions[1].height = 21

    # 设置条件格式
    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[0.7])
    green_fill = PatternFill(bgColor='C0F0D8')
    green_text = Font(color='006000')
    dxf_between = DifferentialStyle(font=green_text, fill=green_fill)
    rule2 = Rule(type='cellIs', operator='between', dxf=dxf_between, formula=[0.5, 0.7])
    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 1, maxrow), rule1)
    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 1, maxrow), rule2)

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet7(inputname, inputdir, outputdir):
    # 获取文件名

    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = '智能屏推送不成功原因' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    # impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
    sheet4_dataframe = impt_xlsx_resource[
        (impt_xlsx_resource.媒体 == 'smart') & (impt_xlsx_resource.套装 == '不分套装') & (
                impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['安装总数',
                '自有智能屏不在线总数',
                '全设备推送成功率',
                '智能屏推送成功率',
                '物业断电',
                '无法进入',
                '换刊受阻',
                '物业拆机',
                '电梯维保',
                '业主破坏',
                '物业装修',
                '设备整洁不符合要求',
                '设备/cf卡被盗',
                '锁损坏',
                '遮挡设备',
                '业主拆机',
                '合同到期',
                '分众欠费',
                '开发通知停机',
                '机型与系统信息不一致',
                '大楼里面找不到这个点位',
                '花屏',
                '黑屏',
                '屏幕定格',
                '4g模块异常',
                'wifi模块异常',
                '有阴影',
                '遥控失灵',
                '不读卡',
                'lcd同步异常',
                '升级失败',
                '自动重启'
                ])
    sheet4_dataframe = sheet4_dataframe.sort_values('安装总数', ascending=False)

    # 重命名字段名
    sheet4_dataframe.rename(columns={'小时统计推送成功且七天平均>19h比率': '可在线监测比'}, inplace=True)

    # 重置列的顺序
    order = ['安装总数',
             '自有智能屏不在线总数',
             '全设备推送成功率',
             '智能屏推送成功率',
             '物业断电',
             '无法进入',
             '电梯维保',
             '物业拆机',
             '业主破坏',
             '物业装修',
             '换刊受阻',
             '设备整洁不符合要求',
             '设备/cf卡被盗',
             '遮挡设备',
             '业主拆机',
             '分众欠费',
             '合同到期',
             '大楼里面找不到这个点位',
             '开发通知停机',
             '机型与系统信息不一致',
             '4g模块异常',
             'wifi模块异常',
             '黑屏',
             '不读卡',
             '锁损坏',
             '升级失败',
             '屏幕定格',
             '遥控失灵',
             '自动重启',
             '花屏',
             '有阴影',
             'lcd同步异常']
    sheet4_dataframe = sheet4_dataframe[order]

    ##返回行数
    row_1 = sheet4_dataframe.shape[0]
    ##返回列数
    column_1 = sheet4_dataframe.shape[1]

    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet4_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=4)
    writer.save()
    writer.close()
    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 6
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'
    number_format1 = '_ * #,##0_ ;_ * -#,##0_ ;_ * "-"??_ ;_ @_ '

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    # 获取长度时为了让中文识别为两个字符，需要控制encode为u8
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        list1 = []
        for cell in col:
            if cell.value is not None:
                list1.append(len(str(cell.value).encode('utf-8')))
        a = max(list1)
        b = get_column_letter(cell.column)
        sheet.column_dimensions[b].width = a + 2

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column in (4, 5):
                    cell.number_format = number_format
                else:
                    cell.number_format = number_format1

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = '智能屏设备非正常原因分析'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A1:AG1')
    sheet.cell(1, 1).alignment = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    sheet.row_dimensions[1].height = 18.75

    fonta2 = Font(name='等线', size=11, bold=False)

    df = sheet4_dataframe
    array = df.columns.values
    array2 = df.iloc[0, :]
    value1 = int(np.sum(array2[4:15], axis=0))
    value2 = int(np.sum(array2[15:20], axis=0))
    value3 = int(np.sum(array2[20:32], axis=0))
    device_df = df[array[20:32]].iloc[0].sort_values(ascending=False)
    prob = '、'.join(device_df[0:5].index)

    sheet[
        'A2'] = '智能屏设备非正常原因说明-' + date_file + '：\n1、智能屏本周因物业及业主原因导致不在线的设备数有' + str(
        value1) + '台。\n2、此外合同到期、运营维护等原因的设备数' + str(value2) + '台。\n3、本周设备故障数' + str(value3) + '台，' + prob[
                                                                                                    0:-1] + '等故障较多。'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A2:AG2')
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    sheet.cell(2, 1).alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 77

    sheet['A3'] = '推送城市'
    sheet['B3'] = '所有安装数'
    sheet['C3'] = '不在线设备数'
    sheet['D3'] = '全部设备推送成功率'
    sheet['E3'] = '智能机推送成功率'
    sheet['J3'] = '智能机推送成功率'
    sheet['K3'] = '可在线监测比'
    sheet['F4'] = '物业及业主原因'
    sheet['Q4'] = '分众自身原因'
    sheet['V4'] = '设备原因'
    sheet['F3'] = '设备非正常原因'

    sheet.merge_cells('F3:AG3')
    sheet.merge_cells('F4:P4')
    sheet.merge_cells('Q4:U4')
    sheet.merge_cells('V4:AG4')
    sheet.merge_cells('A3:A5')
    sheet.merge_cells('B3:B5')
    sheet.merge_cells('C3:C5')
    sheet.merge_cells('D3:D5')
    sheet.merge_cells('E3:E5')

    for col in sheet.iter_cols(min_row=3, max_row=5, min_col=1, max_col=33):
        for cell in col:
            # 单元格字体
            cell.font = font_title
            # 单元格背景颜色
            cell.fill = color_backgrond
            # 单元格对其方式
            cell.alignment = ali
            # 单元格网格线
            cell.border = border

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 9
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 9
    sheet.column_dimensions['H'].width = 9
    sheet.column_dimensions['I'].width = 9
    sheet.column_dimensions['J'].width = 9
    sheet.column_dimensions['K'].width = 9
    sheet.column_dimensions['L'].width = 9
    sheet.column_dimensions['M'].width = 9
    sheet.column_dimensions['N'].width = 9
    sheet.column_dimensions['O'].width = 9
    sheet.column_dimensions['P'].width = 9
    sheet.column_dimensions['Q'].width = 9
    sheet.column_dimensions['R'].width = 9
    sheet.column_dimensions['S'].width = 9
    sheet.column_dimensions['T'].width = 9
    sheet.column_dimensions['U'].width = 9
    sheet.column_dimensions['V'].width = 9
    sheet.column_dimensions['W'].width = 9
    sheet.column_dimensions['X'].width = 9
    sheet.column_dimensions['Y'].width = 9
    sheet.column_dimensions['Z'].width = 9
    sheet.column_dimensions['AA'].width = 9
    sheet.column_dimensions['AB'].width = 9
    sheet.column_dimensions['AC'].width = 9
    sheet.column_dimensions['AD'].width = 9
    sheet.column_dimensions['AE'].width = 9
    sheet.column_dimensions['AF'].width = 9
    sheet.column_dimensions['AG'].width = 9

    first = FormatObject(type='min', val=0)
    second = FormatObject(type='max', val=100)
    data_bar4 = DataBar(cfvo=[first, second], color="FFC030", showValue=None, minLength=None, maxLength=None)
    data_bar5 = DataBar(cfvo=[first, second], color="90D8A8", showValue=None, minLength=None, maxLength=None)
    data_bar6 = DataBar(cfvo=[first, second], color="FF6060", showValue=None, minLength=None, maxLength=None)
    # 将数据条赋给规则
    rule4 = Rule(type='dataBar', dataBar=data_bar4)
    rule5 = Rule(type='dataBar', dataBar=data_bar5)
    rule6 = Rule(type='dataBar', dataBar=data_bar6)

    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 2, maxrow), rule5)
    sheet.conditional_formatting.add('E%s:E%s' % (minrow + 2, maxrow), rule6)

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet6(inputname, inputdir, outputdir):
    # 获取文件名

    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = '智能屏点位推送及回传数据分析' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    # impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
    sheet4_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'smart') & (impt_xlsx_resource.套装 == '不分套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['安装总数',
                '自有智能屏在线总数',
                '自有智能屏不在线总数',
                '自有非智能屏总数',
                '租用总数',
                '推送成功总数',
                '全设备推送成功率',
                '智能屏推送成功率',
                '小时统计推送成功且七天平均>20h比率'
                ])
    sheet4_dataframe = sheet4_dataframe.sort_values('安装总数', ascending=False)

    # 重命名字段名

    sheet4_dataframe.rename(columns={'小时统计推送成功且七天平均>20h比率': '可在线监测比',
                                     '自有智能屏在线总数': '在线',
                                     '自有智能屏不在线总数': '不在线',
                                     '自有非智能屏总数': '非智能机安装数',
                                     '租用总数': '租用屏数目'
                                     }, inplace=True)
    # 重置列的顺序

    order = ['安装总数',
             '在线',
             '不在线',
             '非智能机安装数',
             '租用屏数目',
             '推送成功总数',
             '全设备推送成功率',
             '智能屏推送成功率',
             '可在线监测比']
    sheet4_dataframe = sheet4_dataframe[order]

    ##返回行数
    row_1 = sheet4_dataframe.shape[0]
    ##返回列数
    column_1 = sheet4_dataframe.shape[1]
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet4_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=4)
    writer.save()
    writer.close()
    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 5
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    # 获取长度时为了让中文识别为两个字符，需要控制encode为u8
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        list1 = []
        for cell in col:
            if cell.value is not None:
                list1.append(len(str(cell.value).encode('utf-8')))
        a = max(list1)
        b = get_column_letter(cell.column)
        sheet.column_dimensions[b].width = a + 2

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column in (9, 10, 8):
                    cell.number_format = number_format

    first = FormatObject(type='min', val=0)
    second = FormatObject(type='max', val=100)
    data_bar4 = DataBar(cfvo=[first, second], color="FFC030", showValue=None, minLength=None, maxLength=None)
    data_bar5 = DataBar(cfvo=[first, second], color="90D8A8", showValue=None, minLength=None, maxLength=None)
    data_bar6 = DataBar(cfvo=[first, second], color="FF6060", showValue=None, minLength=None, maxLength=None)
    # 将数据条赋给规则
    rule4 = Rule(type='dataBar', dataBar=data_bar4)
    rule5 = Rule(type='dataBar', dataBar=data_bar5)
    rule6 = Rule(type='dataBar', dataBar=data_bar6)

    sheet.conditional_formatting.add('H%s:H%s' % (minrow + 2, maxrow), rule4)
    sheet.conditional_formatting.add('I%s:I%s' % (minrow + 2, maxrow), rule5)
    sheet.conditional_formatting.add('J%s:J%s' % (minrow + 2, maxrow), rule6)

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = '智能屏推送率及监测率数据分析'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A1:J1')
    sheet.cell(1, 1).alignment = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    sheet.row_dimensions[1].height = 18.75

    fonta2 = Font(name='等线', size=11, bold=False)

    df = sheet4_dataframe
    df = df[df.index.values != '全国']
    city90 = df[df['全设备推送成功率'] > 0.90]
    city90 = '、'.join(city90.index)

    city90_2 = df[df['可在线监测比'] > 0.90]
    city90_2 = '、'.join(city90_2.index)

    sheet[
        'A2'] = '智能屏资源推送率及监测率情况说明-' + date_file + '：\n1、本周整体推送成功率为' + str(
        '%.1f' % (round(sheet['I6'].value * 100, 1))) + '%\n2、本周' + city90 + str(
        np.sum(df['全设备推送成功率'] > 0.90)) + '个城市推送成功率超过90%\n3、本周整体可在线监测比达到' + str(
        '%.1f' % (round(sheet['J6'].value * 100, 1))) + '%；\n4、本周有' + str(
        np.sum(df['可在线监测比'] > sheet['J6'].value)) + '个城市可监测率的设备比例达到' + str(
        '%.1f' % (round(sheet['J6'].value * 100, 1))) + '%，' + str(
        np.sum(df['可在线监测比'] > 0.85)) + '个城市大于85%，' + city90_2 + str(np.sum(df['可在线监测比'] > 0.90)) + '个城市可监测设备比例大于90%。'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A2:J2')
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    sheet.row_dimensions[2].height = 152
    sheet.cell(2, 1).alignment = Alignment(vertical='center', wrapText=True)

    sheet['C4'] = '智能机安装数'
    sheet['E4'] = '非智能机安装数'
    sheet['C3'] = '自有屏幕数'
    sheet['A3'] = '城市'
    sheet['B3'] = '所有安装数'
    sheet['F3'] = '租用屏数目'
    sheet['G3'] = '成功推送设备数'
    sheet['H3'] = '全部设备推送成功率'
    sheet['I3'] = '智能机推送成功率'
    sheet['J3'] = '可在线监测比'

    sheet.merge_cells('C4:D4')
    sheet.merge_cells('E4:E5')
    sheet.merge_cells('C3:E3')
    sheet.merge_cells('A3:A5')
    sheet.merge_cells('B3:B5')
    sheet.merge_cells('F3:F5')
    sheet.merge_cells('G3:G5')
    sheet.merge_cells('H3:H5')
    sheet.merge_cells('I3:I5')
    sheet.merge_cells('J3:J5')

    for col in sheet.iter_cols(min_row=3, max_row=5, min_col=1, max_col=10):
        for cell in col:
            # 单元格字体
            cell.font = font_title
            # 单元格背景颜色
            cell.fill = color_backgrond
            # 单元格对其方式
            cell.alignment = ali
            # 单元格网格线
            cell.border = border
    sheet.row_dimensions[2].height = 129

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 9
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 9
    sheet.column_dimensions['H'].width = 9
    sheet.column_dimensions['I'].width = 9
    sheet.column_dimensions['J'].width = 9

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet5(inputname, inputdir, outputdir):
    # 获取文件名

    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = 'LCD推送不成功原因' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    # impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
    sheet4_dataframe = impt_xlsx_resource[
        (impt_xlsx_resource.媒体 == 'lcd') & (impt_xlsx_resource.套装 == '不分套装') & (
                impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['安装总数',
                '自有智能屏不在线总数',
                '全设备推送成功率',
                '智能屏推送成功率',
                '物业断电',
                '无法进入',
                '换刊受阻',
                '物业拆机',
                '电梯维保',
                '业主破坏',
                '物业装修',
                '设备整洁不符合要求',
                '设备/cf卡被盗',
                '锁损坏',
                '遮挡设备',
                '业主拆机',
                '合同到期',
                '分众欠费',
                '开发通知停机',
                '机型与系统信息不一致',
                '大楼里面找不到这个点位',
                '花屏',
                '黑屏',
                '屏幕定格',
                '4g模块异常',
                'wifi模块异常',
                '有阴影',
                '遥控失灵',
                '不读卡',
                'lcd同步异常',
                '升级失败',
                '自动重启'
                ])
    sheet4_dataframe = sheet4_dataframe.sort_values('安装总数', ascending=False)

    # 重命名字段名
    sheet4_dataframe.rename(columns={'小时统计推送成功且七天平均>19h比率': '可在线监测比'}, inplace=True)

    # 重置列的顺序
    order = ['安装总数',
             '自有智能屏不在线总数',
             '全设备推送成功率',
             '智能屏推送成功率',
             '物业断电',
             '无法进入',
             '换刊受阻',
             '物业拆机',
             '电梯维保',
             '业主破坏',
             '物业装修',
             '设备整洁不符合要求',
             '设备/cf卡被盗',
             '遮挡设备',
             '业主拆机',
             '合同到期',
             '分众欠费',
             '开发通知停机',
             '机型与系统信息不一致',
             '大楼里面找不到这个点位',
             '花屏',
             '黑屏',
             '屏幕定格',
             '4g模块异常',
             'wifi模块异常',
             '有阴影',
             '遥控失灵',
             '不读卡',
             '锁损坏',
             'lcd同步异常',
             '升级失败',
             '自动重启']
    sheet4_dataframe = sheet4_dataframe[order]

    ##返回行数
    row_1 = sheet4_dataframe.shape[0]
    ##返回列数
    column_1 = sheet4_dataframe.shape[1]
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet4_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=4)
    writer.save()
    writer.close()
    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")

    workbook._active_sheet_index = 4
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'
    number_format1 = '_ * #,##0_ ;_ * -#,##0_ ;_ * "-"??_ ;_ @_ '

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    # 获取长度时为了让中文识别为两个字符，需要控制encode为u8
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        list1 = []
        for cell in col:
            if cell.value is not None:
                list1.append(len(str(cell.value).encode('utf-8')))
        a = max(list1)
        b = get_column_letter(cell.column)
        sheet.column_dimensions[b].width = a + 2

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border

            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column in (4, 5):
                    cell.number_format = number_format
                else:
                    cell.number_format = number_format1

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = 'LCD设备非正常原因分析'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A1:AG1')
    sheet.cell(1, 1).alignment = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    sheet.row_dimensions[1].height = 18.75

    fonta2 = Font(name='等线', size=11, bold=False)

    df = sheet4_dataframe
    array = df.columns.values
    array2 = df.iloc[0, :]
    value1 = int(np.sum(array2[4:15], axis=0))
    value2 = int(np.sum(array2[15:20], axis=0))
    value3 = int(np.sum(array2[20:32], axis=0))
    device_df = df[array[20:32]].iloc[0].sort_values(ascending=False)
    prob = '、'.join(device_df[0:5].index)

    sheet[
        'A2'] = 'LCD设备非正常原因说明-' + date_file + '：\n1、LCD本周因物业及业主原因导致不在线的设备数有' + str(
        value1) + '台。\n2、本周分众自身合同到期、运营维护等原因共' + str(value2) + '台。\n3、本周故障设备共' + str(value3) + '台，' + prob[
                                                                                                     0:-1] + '等问题较为严重。'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A2:AG2')
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    sheet.cell(2, 1).alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 77

    sheet['A3'] = '推送城市'
    sheet['B3'] = '所有安装数'
    sheet['C3'] = '不在线设备数'
    sheet['D3'] = '全部设备推送成功率'
    sheet['E3'] = '智能机推送成功率'
    sheet['J3'] = '智能机推送成功率'
    sheet['K3'] = '可在线监测比'
    sheet['F4'] = '物业及业主原因'
    sheet['Q4'] = '分众自身原因'
    sheet['V4'] = '设备原因'
    sheet['F3'] = '设备非正常原因'

    sheet.merge_cells('F3:AG3')
    sheet.merge_cells('F4:P4')
    sheet.merge_cells('Q4:U4')
    sheet.merge_cells('V4:AG4')
    sheet.merge_cells('A3:A5')
    sheet.merge_cells('B3:B5')
    sheet.merge_cells('C3:C5')
    sheet.merge_cells('D3:D5')
    sheet.merge_cells('E3:E5')

    for col in sheet.iter_cols(min_row=3, max_row=5, min_col=1, max_col=33):
        for cell in col:
            # 单元格字体
            cell.font = font_title
            # 单元格背景颜色
            cell.fill = color_backgrond
            # 单元格对其方式
            cell.alignment = ali
            # 单元格网格线
            cell.border = border

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 9
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 9
    sheet.column_dimensions['H'].width = 9
    sheet.column_dimensions['I'].width = 9
    sheet.column_dimensions['J'].width = 9
    sheet.column_dimensions['K'].width = 9
    sheet.column_dimensions['L'].width = 9
    sheet.column_dimensions['M'].width = 9
    sheet.column_dimensions['N'].width = 9
    sheet.column_dimensions['O'].width = 9
    sheet.column_dimensions['P'].width = 9
    sheet.column_dimensions['Q'].width = 9
    sheet.column_dimensions['R'].width = 9
    sheet.column_dimensions['S'].width = 9
    sheet.column_dimensions['T'].width = 9
    sheet.column_dimensions['U'].width = 9
    sheet.column_dimensions['V'].width = 9
    sheet.column_dimensions['W'].width = 9
    sheet.column_dimensions['X'].width = 9
    sheet.column_dimensions['Y'].width = 9
    sheet.column_dimensions['Z'].width = 9
    sheet.column_dimensions['AA'].width = 9
    sheet.column_dimensions['AB'].width = 9
    sheet.column_dimensions['AC'].width = 9
    sheet.column_dimensions['AD'].width = 9
    sheet.column_dimensions['AE'].width = 9
    sheet.column_dimensions['AF'].width = 9
    sheet.column_dimensions['AG'].width = 9

    first = FormatObject(type='min', val=0)
    second = FormatObject(type='max', val=100)
    data_bar4 = DataBar(cfvo=[first, second], color="FFC030", showValue=None, minLength=None, maxLength=None)
    data_bar5 = DataBar(cfvo=[first, second], color="90D8A8", showValue=None, minLength=None, maxLength=None)
    data_bar6 = DataBar(cfvo=[first, second], color="FF6060", showValue=None, minLength=None, maxLength=None)
    # 将数据条赋给规则
    rule4 = Rule(type='dataBar', dataBar=data_bar4)
    rule5 = Rule(type='dataBar', dataBar=data_bar5)
    rule6 = Rule(type='dataBar', dataBar=data_bar6)

    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 2, maxrow), rule5)
    sheet.conditional_formatting.add('E%s:E%s' % (minrow + 2, maxrow), rule6)

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet4(inputname, inputdir, outputdir):
    # 获取文件名

    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = 'LCD分套装点位推送及回传数据分析' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
    sheet4_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'lcd') & (impt_xlsx_resource.套装 != '无套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index=['城市', '套装'],
        values=['安装总数',
                '自有智能屏在线总数',
                '自有智能屏不在线总数',
                '自有非智能屏幕13寸总数',
                '自有非智能屏幕非13寸总数',
                '租用总数',
                '推送成功总数',
                '全设备推送成功率',
                '智能屏推送成功率',
                '小时统计推送成功且七天平均>20h比率'
                ])

    '''
    # 重命名字段名
    sheet4_dataframe.rename(columns={'可售安装总数': '可售点位数',
                                      '推送成功总数': '全部设备推送成功率',
                                      '分钟统计推送成功且最近一天>1h比率': '可在线监测比'}, inplace=True)
    '''

    # 获取各城市套装数
    suit_no = sheet4_dataframe.pivot_table(index='城市', values='安装总数', aggfunc=len)
    suit_no.rename(columns={'安装总数': '套装数'}, inplace=True)

    # 制定套装index
    suit_index = {'套装': ['合计', 'A1套', 'A2套', 'A3套', 'A4套'],
                  'suitindex': [1, 2, 3, 4, 5]
                  }
    suit_index = pd.DataFrame(suit_index)

    # 根据指定index顺序排序
    sheet4_dataframe = sheet4_dataframe.reset_index('套装')
    data_normal = {
        '城市': ['全国', '上海', '北京', '成都', '杭州', '深圳', '广州', '重庆', '天津', '武汉', '昆明', '大连', '南京', '青岛', '合肥', '济南', '长沙',
               '石家庄',
               '长春', '哈尔滨', '沈阳', '厦门', '苏州', '贵阳', '西安', '郑州', '福州', '无锡', '东莞', '太原', '温州', '海口', '烟台', '宁波', '泉州',
               '兰州',
               '佛山', '中山', '珠海', '柳州', '洛阳', '晋江', '绵阳', '廊坊', '漳州', '西宁', '昆山', '常德', '宜昌', '桂林', '惠州', '芜湖', '鄂尔多斯',
               '石狮',
               '江阴', '汕头', '沧州', '常熟', '西双版纳', '三明', '南通', '威海', '张家港', '台州', '泸州', '德阳', '泰州', '临沂', '滁州', '蒙自', '襄阳',
               '衡水', '遵义', '惠安', '绍兴', '义乌', '宜宾', '嘉兴', '安庆', '曲靖', '楚雄', '湖州', '瑞安', '南安', '平湖', '济宁', '玉溪', '湘潭',
               '泰安',
               '三亚', '清镇', '大理', '凯里', '丽江', '九江', '扬州', '安宁', '北碚', '仁怀', '邢台', '个旧', '唐山', '安顺', '咸阳', '都江堰', '新郑',
               '荥阳',
               '乐清', '秦皇岛', '太仓', '慈溪', '巢湖', '涪陵', '永川', '邯郸', '保定'],
        'no': [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
               29,
               30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55,
               56,
               57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82,
               83,
               84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105, 106, 107,
               108,
               109, 110, 111, 112, 113, 114, 115]}
    data_normal = pd.DataFrame(data_normal)

    # 融合多表并去除单套合计数据
    merge_1 = pd.merge(sheet4_dataframe, data_normal, left_on='城市', right_on='城市', how='inner', sort=False)
    merge_2 = pd.merge(merge_1, suit_no, left_on='城市', right_on='城市', how='inner', sort=False)
    merge_3 = pd.merge(merge_2, suit_index, left_on='套装', right_on='套装', how='inner', sort=False)
    merge_3 = merge_3[(merge_3['套装数'] > 2) | ((merge_3['套装数'] == 2) & (merge_3['套装'] != '合计')) | (
            (merge_3['套装数'] == 1) & (merge_3['套装'] == '合计'))]

    # 对融合后的表处理，保留所需结果字段
    merge_result = merge_3.pivot_table(index=['城市', '套装'],
                                       values=['安装总数',
                                               '自有智能屏在线总数',
                                               '自有智能屏不在线总数',
                                               '自有非智能屏幕13寸总数',
                                               '自有非智能屏幕非13寸总数',
                                               '租用总数',
                                               '推送成功总数',
                                               '全设备推送成功率',
                                               '智能屏推送成功率',
                                               '小时统计推送成功且七天平均>20h比率',
                                               'no',
                                               'suitindex'])
    merge_result = merge_result.sort_values(by=['no', 'suitindex'])
    merge_result.drop(['no', 'suitindex'], axis=1, inplace=True)

    merge_result.rename(columns={'小时统计推送成功且七天平均>20h比率': '可在线监测比',
                                 '自有智能屏在线总数': '在线',
                                 '自有智能屏不在线总数': '不在线',
                                 '自有非智能屏幕13寸总数': '13寸',
                                 '自有非智能屏幕非13寸总数': '非13寸'
                                 }, inplace=True)
    # 重置列的顺序
    order = ['安装总数',
             '在线',
             '不在线',
             '13寸',
             '非13寸',
             '租用总数',
             '推送成功总数',
             '全设备推送成功率',
             '智能屏推送成功率',
             '可在线监测比']
    merge_result = merge_result[order]

    ##返回行数
    row_1 = merge_result.shape[0]
    ##返回列数
    column_1 = merge_result.shape[1]
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    merge_result.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=4)
    writer.save()
    writer.close()

    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 3
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0%'

    # 循环每一列，获取对应每一列中字符串最大长度，并对这一列的宽度做更改
    # 获取长度时为了让中文识别为两个字符，需要控制encode为u8
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        list1 = []
        for cell in col:
            if cell.value is not None:
                list1.append(len(str(cell.value).encode('utf-8')))
        a = max(list1)
        b = get_column_letter(cell.column)
        sheet.column_dimensions[b].width = a + 2

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column in (10, 11, 12):
                    cell.number_format = number_format

    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[0.95])
    green_fill = PatternFill(bgColor='C0F0D8')
    green_text = Font(color='006000')
    dxf_between = DifferentialStyle(font=green_text, fill=green_fill)
    rule2 = Rule(type='cellIs', operator='between', dxf=dxf_between, formula=[0.9, 0.95])
    yellow_fill = PatternFill(bgColor='FFF090')
    yellow_text = Font(color='906000')
    dxf_yellow = DifferentialStyle(font=yellow_text, fill=yellow_fill)
    rule3 = Rule(type='cellIs', operator='between', dxf=dxf_yellow, formula=[0.9, 0.95])

    sheet.conditional_formatting.add('J%s:L%s' % (minrow + 1, maxrow), rule1)
    sheet.conditional_formatting.add('J%s:J%s' % (minrow + 1, maxrow), rule3)

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = 'LCD分套装推送率及监测率数据分析'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A1:L1')
    sheet.cell(1, 1).alignment = Alignment(horizontal='centerContinuous', vertical='justify', wrapText=True)
    sheet.row_dimensions[1].height = 18.75

    fonta2 = Font(name='等线', size=11, bold=False)
    df = merge_result.reset_index('套装')
    total = df[df['套装'] != '合计']
    df1 = total['全设备推送成功率']
    df2 = total['可在线监测比']
    sheet[
        'A2'] = 'LCD分套装资源推送率及监测率情况说明-' + date_file + '：\n1、本周全国整体推送率' + str(
        '%.1f' % (round(sheet['J6'].value * 100, 1))) + '%；\n2、LCD当前共' + str(
        np.sum(df1 > 0.90)) + '个套装的推送率超过90%，其中，' + str(np.sum(df1 > 0.95)) + '个套装的推送成功率超过95%。\n3、本周整体可在线监测的比例达到' + str(
        '%.1f' % (round(sheet['L6'].value * 100, 1))) + '%；\n4、LCD当前' + str(
        np.sum(df2 > 0.90)) + '个套装可监测设备占比超过90%，其中，' + str(np.sum(df2 > 0.95)) + '个套装可监测比例超过95％。'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A2:L2')
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    sheet.row_dimensions[2].height = 152
    sheet.cell(2, 1).alignment = Alignment(vertical='center', wrapText=True)

    #sheet['D4'] = '智能机安装数'
    #sheet['F4'] = '非智能机安装数'
    sheet['D3'] = '自有屏幕数'

    sheet['A3'] = '城市'
    sheet['B3'] = '套装'
    sheet['C3'] = '所有安装数'
    sheet['H3'] = '租用屏数目'
    sheet['I3'] = '成功推送设备数'
    sheet['J3'] = '全部设备推送成功率'
    sheet['K3'] = '智能机推送成功率'
    sheet['L3'] = '可在线监测比'

    sheet['D4'] = order[1]
    sheet['E4'] = order[2]
    sheet['F4'] = order[3]
    sheet['G4'] = order[4]

    #sheet.merge_cells('D4:E4')
    #sheet.merge_cells('F4:G4')
    sheet.merge_cells('D3:G3')


    sheet.merge_cells('D4:D5')
    sheet.merge_cells('E4:E5')
    sheet.merge_cells('F4:F5')
    sheet.merge_cells('G4:G5')

    sheet.merge_cells('A3:A5')
    sheet.merge_cells('B3:B5')
    sheet.merge_cells('C3:C5')
    sheet.merge_cells('H3:H5')
    sheet.merge_cells('I3:I5')
    sheet.merge_cells('J3:J5')
    sheet.merge_cells('K3:K5')
    sheet.merge_cells('L3:L5')

    for col in sheet.iter_cols(min_row=3, max_row=5, min_col=1, max_col=12):
        for cell in col:
            # 单元格字体
            cell.font = font_title
            # 单元格背景颜色
            cell.fill = color_backgrond
            # 单元格对其方式
            cell.alignment = ali
            # 单元格网格线
            cell.border = border
    sheet.row_dimensions[2].height = 96

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 9
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 9
    sheet.column_dimensions['H'].width = 9
    sheet.column_dimensions['I'].width = 9
    sheet.column_dimensions['J'].width = 10
    sheet.column_dimensions['K'].width = 10
    sheet.column_dimensions['L'].width = 10

    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet3(inputname, inputdir, outputdir):
    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = 'LCD点位推送及回传数据分析' + date_file
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    # impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
    sheet4_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'lcd') & (impt_xlsx_resource.套装 == '不分套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['安装总数',
                '自有智能屏在线总数',
                '自有智能屏不在线总数',
                '自有非智能屏幕13寸总数',
                '自有非智能屏幕非13寸总数',
                '租用总数',
                '推送成功总数',
                '全设备推送成功率',
                '智能屏推送成功率',
                '小时统计推送成功且七天平均>20h比率'
                ])
    sheet4_dataframe = sheet4_dataframe.sort_values('安装总数', ascending=False)

    # 重命名字段名
    sheet4_dataframe.rename(columns={'小时统计推送成功且七天平均>20h比率': '可在线监测比',
                                     '自有智能屏在线总数': '在线',
                                     '自有智能屏不在线总数': '不在线',
                                     '自有非智能屏幕13寸总数': '13寸',
                                     '自有非智能屏幕非13寸总数': '非13寸'
                                     }, inplace=True)

    # 重置列的顺序
    order = ['安装总数',
             '在线',
             '不在线',
             '13寸',
             '非13寸',
             '租用总数',
             '推送成功总数',
             '全设备推送成功率',
             '智能屏推送成功率',
             '可在线监测比']
    sheet4_dataframe = sheet4_dataframe[order]

    ##返回行数
    row_1 = sheet4_dataframe.shape[0]
    ##返回列数
    column_1 = sheet4_dataframe.shape[1]
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet4_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=4)
    writer.save()
    writer.close()

    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 2
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0.0%'

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column in (9, 10, 11):
                    cell.number_format = number_format

    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='top10', operator='greaterThan', percent=True, dxf=dxf, rank=10)
    rule2 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[100])
    rule3 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[500])

    first = FormatObject(type='min', val=0)
    second = FormatObject(type='max', val=100)
    data_bar4 = DataBar(cfvo=[first, second], color="FFC030", showValue=None, minLength=None, maxLength=None)
    data_bar5 = DataBar(cfvo=[first, second], color="90D8A8", showValue=None, minLength=None, maxLength=None)
    data_bar6 = DataBar(cfvo=[first, second], color="FF6060", showValue=None, minLength=None, maxLength=None)
    # 将数据条赋给规则
    rule4 = Rule(type='dataBar', dataBar=data_bar4)
    rule5 = Rule(type='dataBar', dataBar=data_bar5)
    rule6 = Rule(type='dataBar', dataBar=data_bar6)

    sheet.conditional_formatting.add('D%s:D%s' % (minrow + 2, maxrow), rule1)
    sheet.conditional_formatting.add('E%s:E%s' % (minrow + 2, maxrow), rule1)
    sheet.conditional_formatting.add('F%s:F%s' % (minrow + 2, maxrow), rule2)
    sheet.conditional_formatting.add('G%s:G%s' % (minrow + 2, maxrow), rule3)
    sheet.conditional_formatting.add('I%s:I%s' % (minrow + 2, maxrow), rule4)
    sheet.conditional_formatting.add('J%s:J%s' % (minrow + 2, maxrow), rule5)
    sheet.conditional_formatting.add('K%s:K%s' % (minrow + 2, maxrow), rule6)

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = 'LCD点位推送率及监测率数据分析'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A1:K1')
    sheet.cell(1, 1).alignment = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    sheet.row_dimensions[1].height = 18.75
    fonta2 = Font(name='等线', size=11, bold=False)
    # 找占比前四， （租用屏数/所有安装数）当影响较大
    df = sheet4_dataframe
    df = df[df.index.values != '全国']
    df['占比'] = df['租用总数'] / df['安装总数']
    df = df.sort_values(by='占比', ascending=0)
    total = df.index.values
    total2 = df[df['全设备推送成功率'] > 0.9]
    total3 = df[df['可在线监测比'] > 0.90]
    if len(total) > 4:
        city_zu = '、'.join(total[0:4])
    else:
        city_zu = '、'.join(total)
    if len(total2) > 22:
        city_success = '、'.join(total2[0:22].index)
    else:
        city_success = '、'.join(total2.index)
    if len(total3) > 4:
        city_online = '、'.join(total3[0:4].index)
    else:
        city_online = '、'.join(total3.index)

    sheet['A2'] = 'LCD资源推送率及监测率情况说明-' + date_file + '：\n1、本周LCD资源推送率达到' + str(
        '%.1f' % (round(sheet['I6'].value * 100, 1))) + '%；\n2、不能推送的租用屏占比' + str('%.2f' % (
        round((sheet['G6'].value / sheet['B6'].value) * 100, 2))) + '%，影响较大的城市包括' + city_zu + '等。\n3、全国有' + str(
        np.sum(df['全设备推送成功率'] > 0.80)) + '个城市推送成功率超过80%，其中' + city_success + '等' + str(
        np.sum(df['全设备推送成功率'] > 0.90)) + '个城市点位推送率已达到90%以上。\n4、本周LCD资源总体可在线监测比达到' + str(
        '%.1f' % (round(sheet['K6'].value * 100, 1))) + '%；\n5、可监播设备比例超过70%的有' + str(
        np.sum(df['可在线监测比'] > 0.70)) + '个城市，超过80%的有' + str(
        np.sum(df['可在线监测比'] > 0.80)) + '个城市，其中' + city_online + '等' + str(np.sum(df['可在线监测比'] > 0.90)) + '个城市可达到90%以上。'
    sheet['A1'].font = fonta1
    sheet.merge_cells('A2:K2')
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    sheet.cell(2, 1).alignment = Alignment(vertical='center', wrapText=True)

    #sheet['C4'] = '智能机安装数'
    #sheet['E4'] = '非智能机安装数'
    sheet['C3'] = '自有屏幕数'
    sheet['A3'] = '城市'
    sheet['B3'] = '所有安装数'
    sheet['G3'] = '租用屏数目'
    sheet['H3'] = '成功推送设备数'
    sheet['I3'] = '全部设备推送成功率'
    sheet['J3'] = '智能机推送成功率'
    sheet['K3'] = '可在线监测比'

    sheet['C4'] = order[1]
    sheet['D4'] = order[2]
    sheet['E4'] = order[3]
    sheet['F4'] = order[4]

    #sheet.merge_cells('C4:D4')
    #sheet.merge_cells('E4:F4')
    sheet.merge_cells('C3:F3')

    sheet.merge_cells('C4:C5')
    sheet.merge_cells('D4:D5')
    sheet.merge_cells('E4:E5')
    sheet.merge_cells('F4:F5')

    sheet.merge_cells('A3:A5')
    sheet.merge_cells('B3:B5')
    sheet.merge_cells('G3:G5')
    sheet.merge_cells('H3:H5')
    sheet.merge_cells('I3:I5')
    sheet.merge_cells('J3:J5')
    sheet.merge_cells('K3:K5')

    for col in sheet.iter_cols(min_row=3, max_row=5, min_col=1, max_col=11):
        for cell in col:
            # 单元格字体
            cell.font = font_title
            # 单元格背景颜色
            cell.fill = color_backgrond
            # 单元格对其方式
            cell.alignment = ali
            # 单元格网格线
            cell.border = border
    sheet.row_dimensions[2].height = 152
    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet2(inputname, inputdir, outputdir):
    # 获取文件名
    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = '智能屏点位推送及回传数据分析' + date_file + '销售用'
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    sheet2_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'smart') & (impt_xlsx_resource.套装 == '不分套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index='城市',
        values=['可售安装总数', '全设备推送成功率', '小时统计推送成功且七天平均>20h比率'],
        aggfunc={'可售安装总数': np.sum, '全设备推送成功率': np.sum, '小时统计推送成功且七天平均>20h比率': np.sum})
    sheet2_dataframe = sheet2_dataframe.sort_values('可售安装总数', ascending=False)

    # 重命名字段名
    sheet2_dataframe.rename(columns={'可售安装总数': '可售点位数',
                                     '全设备推送成功率': '全部设备推送成功率',
                                     '小时统计推送成功且七天平均>20h比率': '可在线监测比'}, inplace=True)

    # 重置列的顺序
    order = ['可售点位数', '全部设备推送成功率', '可在线监测比']
    sheet2_dataframe = sheet2_dataframe[order]

    ##返回行数
    row_1 = sheet2_dataframe.shape[0]
    ##返回列数
    column_1 = sheet2_dataframe.shape[1] - 1
    writer = pd.ExcelWriter(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheet2_dataframe.to_excel(excel_writer=writer, sheet_name=result_sheetname, startrow=2)
    writer.save()
    writer.close()

    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    workbook._active_sheet_index = 1
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0.0%'
    sheet.row_dimensions[2].height = 120
    sheet.row_dimensions[1].height = 21
    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = '智能屏资源推送率及监测率数据分析'
    sheet['A1'].font = fonta1
    sheet['A1'].alignment = Alignment(wrapText=True)
    sheet['A1'].alignment = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    sheet.merge_cells('A1:D1')

    df = sheet2_dataframe
    df = df[df.index.values != '全国']
    total = df[df['全部设备推送成功率'] > 0.90]
    city90 = '、'.join(total.index)
    df2 = df['可在线监测比']
    average = str('%.1f' % (round(sheet['D4'].value * 100, 1))) + '%'
    total2 = df[df2 > 0.90]
    city90_2 = '、'.join(total2.index)

    sheet[
        'A2'] = '''智能屏资源推送率及监测率情况说明-''' + date_file + '''：\n1、本周整体推送成功率为''' + str(
        '%.1f' % (round(sheet['C4'].value * 100, 1))) + '''%;\n2、本周''' + city90 + str(
        np.sum(df['全部设备推送成功率'] > 0.90)) + '''个城市推送成功率超过90%\n3、本周整体可在线监测比达到''' + average + ''';\n4、本周有''' + str(
        np.sum(df['可在线监测比'] > sheet['D4'].value)) + '''个城市可监测率的设备比例达到''' + average + '''，''' + str(
        np.sum(df['可在线监测比'] > 0.85)) + '''个城市大于85%，''' + city90_2 + str(np.sum(df2 > 0.90)) + '''个城市可监测设备比例大于90%'''
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    sheet.merge_cells('A2:D2')
    sheet['A2'].alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 128.5
    sheet.row_dimensions[1].height = 21

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column == 3 or cell.column == 4:
                    cell.number_format = number_format

    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[0.95])
    green_fill = PatternFill(bgColor='C0F0D8')
    green_text = Font(color='006000')
    dxf_between = DifferentialStyle(font=green_text, fill=green_fill)
    rule2 = Rule(type='cellIs', operator='between', dxf=dxf_between, formula=[0.9, 0.95])
    sheet.conditional_formatting.add('C%s:D%s' % (minrow + 1, maxrow), rule1)
    sheet.conditional_formatting.add('C%s:D%s' % (minrow + 1, maxrow), rule2)

    sheet.column_dimensions['A'].width = 15.62
    sheet.column_dimensions['B'].width = 15.62
    sheet.column_dimensions['C'].width = 15.62
    sheet.column_dimensions['D'].width = 15.62
    # 关闭默认灰色网格线
    sheet.sheet_view.showGridLines = False
    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


def monitor_result_sheet1(inputname, inputdir, outputdir):
    # 获取文件名
    # date_file = inputname[len(inputname) - 8:len(inputname) - 4]
    # 确定结果sheet名
    result_sheetname = 'LCD分套装点位推送及回传数据分析' + date_file + '销售用'
    impt_xlsx_resource = pd.read_csv(inputdir + inputname, sep='\t')
    impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
    sheet4_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'lcd') & (impt_xlsx_resource.套装 != '无套装') & (
            impt_xlsx_resource.城市 != '无匹配')].pivot_table(
        index=['城市', '套装'],
        values=['可售安装总数', '全设备推送成功率', '小时统计推送成功且七天平均>20h比率'])

    # 重命名字段名
    sheet4_dataframe.rename(columns={'可售安装总数': '可售点位数',
                                     '全设备推送成功率': '全部设备推送成功率',
                                     '小时统计推送成功且七天平均>20h比率': '可在线监测比'}, inplace=True)
    # 获取各城市套装数
    suit_no = sheet4_dataframe.pivot_table(index='城市', values='可售点位数', aggfunc=len)
    suit_no.rename(columns={'可售点位数': '套装数'}, inplace=True)

    # 制定套装index
    suit_index = {'套装': ['合计', 'A1套', 'A2套', 'A3套', 'A4套'],
                  'suitindex': [1, 2, 3, 4, 5]
                  }
    suit_index = pd.DataFrame(suit_index)

    # 根据指定index顺序排序
    sheet4_dataframe = sheet4_dataframe.reset_index('套装')
    data_normal = {
        '城市': ['全国', '上海', '北京', '成都', '杭州', '深圳', '广州', '重庆', '天津', '武汉', '昆明', '大连', '南京', '青岛', '合肥', '济南', '长沙',
               '石家庄', '长春', '哈尔滨', '沈阳', '厦门', '苏州', '贵阳', '西安', '郑州', '福州', '无锡', '东莞', '太原', '温州', '海口', '烟台', '宁波',
               '泉州', '兰州', '佛山', '中山', '珠海', '柳州', '洛阳', '晋江', '绵阳', '廊坊', '漳州', '西宁', '昆山', '常德', '宜昌', '桂林', '惠州',
               '芜湖', '鄂尔多斯', '石狮', '江阴', '汕头', '沧州', '常熟', '西双版纳', '三明', '南通', '威海', '张家港', '台州', '泸州', '德阳', '泰州',
               '临沂', '滁州', '蒙自', '襄阳', '衡水', '遵义', '惠安', '绍兴', '义乌', '宜宾', '嘉兴', '安庆', '曲靖', '楚雄', '湖州', '瑞安', '南安',
               '平湖', '济宁', '玉溪', '湘潭', '泰安', '三亚', '清镇', '大理', '凯里', '丽江', '九江', '扬州', '安宁', '北碚', '仁怀', '邢台', '个旧',
               '唐山', '安顺', '咸阳', '都江堰', '新郑', '荥阳', '乐清', '秦皇岛', '太仓', '慈溪', '巢湖', '涪陵', '永川', '邯郸', '保定'],
        'no': [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
               29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54,
               55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80,
               81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105,
               106, 107, 108, 109, 110, 111, 112, 113, 114, 115]}
    data_normal = pd.DataFrame(data_normal)

    # 融合多表并去除单套合计数据
    merge_1 = pd.merge(sheet4_dataframe, data_normal, left_on='城市', right_on='城市', how='inner', sort=False)
    merge_2 = pd.merge(merge_1, suit_no, left_on='城市', right_on='城市', how='inner', sort=False)
    merge_3 = pd.merge(merge_2, suit_index, left_on='套装', right_on='套装', how='inner', sort=False)
    merge_3 = merge_3[(merge_3['套装数'] > 2) | ((merge_3['套装数'] == 2) & (merge_3['套装'] != '合计')) | (
            (merge_3['套装数'] == 1) & (merge_3['套装'] == '合计'))]

    # 对融合后的表处理，保留所需结果字段
    merge_result = merge_3.pivot_table(index=['城市', '套装'], values=['可售点位数', '全部设备推送成功率', '可在线监测比', 'no', 'suitindex'])
    merge_result = merge_result.sort_values(by=['no', 'suitindex'])

    # merge_result['可在线监测比'] = merge_result['可在线监测比'].apply(lambda x: format(x, '.2%'))
    # merge_result['全部设备推送成功率'] = merge_result['全部设备推送成功率'].apply(lambda x: format(x, '.2%'))
    merge_result.drop(['no', 'suitindex'], axis=1, inplace=True)

    # 重置列的顺序
    order = ['可售点位数', '全部设备推送成功率', '可在线监测比']
    merge_result = merge_result[order]

    ##返回行数
    row_1 = merge_result.shape[0]
    ##返回列数
    column_1 = merge_result.shape[1]

    merge_result.to_excel(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx', sheet_name=result_sheetname, startrow=2)

    ##修改样式,遍历A-Z列
    workbook = load_workbook(outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    ali = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    font = Font(name='等线', size=11, bold=False, italic=False)
    border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                    right=Side(border_style='thin'))
    color_backgrond = PatternFill(fill_type="solid", fgColor="BBBBBB")
    sheet = workbook.active
    first_change = 1
    font_title = Font(name='等线', size=11, bold=True, italic=False)
    maxrow = sheet.max_row
    minrow = sheet.min_row
    maxcol = sheet.max_column
    mincol = sheet.min_column
    dimension = sheet.dimensions
    number_format = '0.0%'

    sheet.column_dimensions['A'].width = 15.62
    sheet.column_dimensions['B'].width = 10.62
    sheet.column_dimensions['C'].width = 15.62
    sheet.column_dimensions['D'].width = 15.62
    sheet.column_dimensions['E'].width = 15.62

    # 遍历选定区域单元格并为首行守列设定特殊标题格式
    for col in sheet.iter_cols(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
        for cell in col:
            if cell.row == minrow or cell.column == mincol or cell.column == mincol + 1:
                # 单元格字体
                cell.font = font_title
                # 单元格背景颜色
                cell.fill = color_backgrond
                # 单元格对其方式
                cell.alignment = ali
                # 单元格网格线
                cell.border = border
            else:
                cell.font = font
                cell.alignment = ali
                cell.border = border
                if cell.column == 4 or cell.column == 5:
                    cell.number_format = number_format

    # 设置条件格式
    red_fill = PatternFill(bgColor='FFC7CE')
    red_text = Font(color='900000')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule1 = Rule(type='cellIs', operator='greaterThan', dxf=dxf, formula=[0.95])
    green_fill = PatternFill(bgColor='C0F0D8')
    green_text = Font(color='006000')
    dxf_between = DifferentialStyle(font=green_text, fill=green_fill)
    rule2 = Rule(type='cellIs', operator='between', dxf=dxf_between, formula=[0.9, 0.95])
    sheet.conditional_formatting.add('D%s:F%s' % (minrow + 1, maxrow), rule1)
    sheet.conditional_formatting.add('D%s:F%s' % (minrow + 1, maxrow), rule2)
    # 关闭默认灰色网格线
    sheet.sheet_view.showGridLines = False

    fonta1 = Font(name='等线', size=14, bold=True)
    sheet['A1'] = 'LCD分套装资源推送率及监测率数据分析'
    sheet['A1'].font = fonta1

    sheet['A1'].alignment = Alignment(horizontal='centerContinuous', vertical='center', wrapText=True)
    sheet.merge_cells('A1:E1')
    # seriesObj = empDfObj.apply(lambda x: True if x['Age'] > 30 else False , axis=1)
    df = merge_result.reset_index('套装')
    total = df[df['套装'] != '合计']
    df1 = total['全部设备推送成功率']
    df2 = total['可在线监测比']
    sheet[
        'A2'] = '''LCD分套装资源推送率及监测率情况说明-''' + date_file + '''：\n1、本周全国整体推送率''' + str(
        '%.1f' % (round(sheet['D4'].value * 100, 1))) + '%' + '''；\n2、LCD当前共''' + str(
        np.sum(df1 > 0.90)) + '''个套装的推送率超过90%，其中，''' + str(
        np.sum(df1 > 0.95)) + '''个套装的推送成功率超过95%。\n3、本周整体可在线监测的比例达到''' + str(
        '%.1f' % (round(sheet['E4'].value * 100, 1))) + '%' + '''；\n4、LCD当前''' + str(
        np.sum(df2 > 0.90)) + '''个套装可监测设备占比超过90%，其中，''' + str(np.sum(df2 > 0.95)) + '''个套装可监测比例超过95％。'''
    fonta2 = Font(name='等线', size=11)
    sheet['A2'].font = fonta2
    # 当通过指定单元格写入带有换行符的文字时，为了让文件打开时显示的就是已换行的，需要对Alignment中的wrapText属性设置为True

    sheet.merge_cells('A2:E2')
    sheet['A2'].alignment = Alignment(vertical='center', wrapText=True)
    sheet.row_dimensions[2].height = 99
    sheet.row_dimensions[1].height = 21

    # 改变格式后的文件反写会对应sheet
    workbook.save(filename=outputdir + 'monitor_reuslt_' + date_file_output + '.xlsx')
    print(result_sheetname + ' Success !')


for root, dir, files in os.walk(inputdir):
    print(root)
    print(files)
    for filez in files:
        if filez.endswith('txt') and filez != 'device.txt' or filez.endswith('csv'):
            inputname = filez
            date_file_output = inputname.split('.')[0]
            monitor_result_sheet1(filez, inputdir, inputdir)
            monitor_result_sheet2(filez, inputdir, inputdir)
            monitor_result_sheet3(filez, inputdir, inputdir)
            monitor_result_sheet4(filez, inputdir, inputdir)
            monitor_result_sheet5(filez, inputdir, inputdir)
            monitor_result_sheet6(filez, inputdir, inputdir)
            monitor_result_sheet7(filez, inputdir, inputdir)
            monitor_result_sheet8(filez, inputdir, inputdir)
            monitor_result_sheet9(filez, inputdir, inputdir)
            monitor_result_sheet10(filez, inputdir, inputdir)
"""
monitor_result_sheet1(inputname, inputdir, outputdir)
monitor_result_sheet2(inputname, inputdir, outputdir)
monitor_result_sheet3(inputname, inputdir, outputdir)
monitor_result_sheet4(inputname, inputdir, outputdir)
monitor_result_sheet5(inputname, inputdir, outputdir)
monitor_result_sheet6(inputname, inputdir, outputdir)
monitor_result_sheet7(inputname, inputdir, outputdir)
monitor_result_sheet8(inputname, inputdir, outputdir)
monitor_result_sheet9(inputname, inputdir, outputdir)
monitor_result_sheet10(inputname, inputdir, outputdir)
"""

